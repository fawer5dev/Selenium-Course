import openpyxl

class Funexcel():
    def __init__(self, driver):
        self.driver = driver

    # Get the number of rows in the Excel sheet
    def getRowCount(self, path, sheetName):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.max_row

    # Get the number of columns in the Excel sheet
    def getColumnCount(self, file, sheetName):
        Workbook = openpyxl.load_workbook(file)
        sheet = Workbook[sheetName]
        return sheet.max_column

    # Read data from the Excel sheet based on row and column number
    def readData(self, path, sheetName, rownum, columnno):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    # Write data to the Excel sheet based on row and column number
    def writeData(self, path, sheetName, rownum, columnno, data):
        Workbook = openpyxl.load_workbook(path)
        sheet = Workbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        Workbook.save(path)
